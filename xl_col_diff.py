import argparse
import asyncio
import os
import re
import openpyxl
import logging

# Configure logging (if not already configured)
logging.basicConfig(
    level=logging.INFO, format="%(asctime)s - %(levelname)s - %(message)s"
)


async def compare_and_sync_columns(
    old_file, new_file, allow_delete, sync_type, ignore_sheet_regex=None
):
    """
    Compares and synchronizes columns between two Excel files.

    Args:
        old_file (str): Path to the old Excel file.
        new_file (str): Path to the new Excel file.
        allow_delete (bool): Whether to allow deletion of columns in the old file.
        sync_type (str): 'rightmost' or 'moverows' to specify how to add new columns.
        ignore_sheet_regex (str, optional): Regex pattern to ignore sheets. Defaults to None.
    """

    if sync_type not in ("rightmost", "moverows"):
        raise ValueError(
            "Invalid sync_type. Must be 'rightmost' or 'moverows'."
        )

    try:
        old_wb = openpyxl.load_workbook(old_file)
        new_wb = openpyxl.load_workbook(new_file)

        for sheet_name in new_wb.sheetnames:
            # Check if the sheet should be ignored
            if ignore_sheet_regex and re.match(ignore_sheet_regex, sheet_name):
                logging.info(f"Ignoring sheet: {sheet_name}")
                continue

            if sheet_name not in old_wb.sheetnames:
                logging.warning(
                    f"Sheet '{sheet_name}' not found in old file. Skipping."
                )
                continue

            old_sheet = old_wb[sheet_name]
            new_sheet = new_wb[sheet_name]

            # Get headers from the first row
            old_headers = [
                cell.value for cell in old_sheet[1]
            ]  # row index starts from 1
            new_headers = [cell.value for cell in new_sheet[1]]

            logging.info(f"Processing sheet: {sheet_name}")
            logging.info(f"Old headers: {old_headers}")
            logging.info(f"New headers: {new_headers}")

            # Add missing columns to old_sheet
            for col_num, header in enumerate(new_headers, start=1):
                if header not in old_headers:
                    if sync_type == "rightmost":
                        # Add to the rightmost column
                        insert_position = old_sheet.max_column + 1
                        old_sheet.cell(row=1, column=insert_position, value=header)
                        old_headers.append(header)
                        logging.info(
                            f"Added column '{header}' to sheet '{sheet_name}' at the rightmost"
                        )

                    elif sync_type == "moverows":
                        # Insert column and shift rows
                        insert_position = col_num
                        old_sheet.insert_cols(insert_position)
                        old_sheet.cell(row=1, column=insert_position, value=header)
                        old_headers.insert(insert_position - 1, header)
                        logging.info(
                            f"Added column '{header}' to sheet '{sheet_name}' at position {insert_position} (shifting rows)"
                        )

            # Delete columns from old_sheet if allow_delete is True
            if allow_delete:
                cols_to_delete = []
                for col_num, header in enumerate(old_headers, start=1):
                    if header not in new_headers:
                        cols_to_delete.append(col_num)

                # Delete columns in reverse order to avoid index shifting
                for col_num in sorted(cols_to_delete, reverse=True):
                    old_sheet.delete_cols(col_num)
                    logging.warning(
                        f"Deleted column {old_headers[col_num - 1]} from sheet '{sheet_name}'"
                    )
                # Refresh old_headers after deletion
                old_headers = [
                    cell.value for cell in old_sheet[1]
                ]  # row index starts from 1

            # Rename columns in old_sheet to match new_sheet
            for col_num, header in enumerate(new_headers, start=1):
                if (
                    col_num <= len(old_headers)
                ):  # Check if the column exists in the old sheet
                    if old_headers[col_num - 1] != header:
                        old_sheet.cell(row=1, column=col_num, value=header)
                        logging.info(
                            f"Renamed column '{old_headers[col_num - 1]}' to '{header}' in sheet '{sheet_name}'"
                        )

        old_wb.save(old_file)
        logging.info(f"Successfully synchronized columns in file: {old_file}")

    except FileNotFoundError:
        logging.error(f"File not found: {old_file} or {new_file}")
    except Exception as e:
        logging.exception(f"An error occurred: {e}")


# Function to traverse directories and compare files
def compare_directory_files(
    old_dir,
    new_dir,
    allow_delete=False,
    sync_type=None,
    ignore_file_regex=None,
    ignore_sheet_regex=None,
):
    for root, dirs, files in os.walk(new_dir):
        for file_name in files:
            if ignore_file_regex and re.match(ignore_file_regex, file_name):
                print(f"Ignoring file: {file_name} (matches ignore regex)")
                continue  # Skip this file if it matches the ignore regex

            if file_name.endswith(('.xlsx', '.xls')):  # Check if it's an Excel file
                # Get relative file path
                relative_path = os.path.relpath(os.path.join(root, file_name), new_dir)
                old_file_path = os.path.join(old_dir, relative_path)

                # Check if the file exists in both directories
                if os.path.exists(old_file_path):
                    print(
                        f"Comparing: {old_file_path} and {os.path.join(root, file_name)}"
                    )
                    asyncio.run(
                        compare_and_sync_columns(
                            old_file_path,
                            os.path.join(root, file_name),
                            allow_delete,
                            sync_type,
                            ignore_sheet_regex,
                        )
                    )


# Main function to parse arguments and execute the script
def main():
    parser = argparse.ArgumentParser(
        description="Compare and sync columns in two Excel files or directories. Updates old sheets with new header names."
    )
    parser.add_argument(
        "--old-file", required=True, help="Path to the old Excel file or directory."
    )
    parser.add_argument(
        "--new-file", required=True, help="Path to the new Excel file or directory."
    )
    parser.add_argument(
        "--check-directory",
        action="store_true",
        help="Specify if the paths are directories to compare Excel files inside.",
    )
    parser.add_argument(
        "--allow-delete",
        action="store_true",
        help="Allow deletion of columns in the old file.",
    )
    parser.add_argument(
        "--sync-type",
        choices=["rightmost", "moverows"],
        default="rightmost",
        help="How to sync columns: 'rightmost' (add to end) or 'moverows' (insert and shift)",
    )
    parser.add_argument(
        "--ignore-sheet-regex",
        type=str,
        help="Regex pattern to ignore sheets during comparison.",
    )
    parser.add_argument(
        "--ignore-file-regex",
        type=str,
        help="Regex pattern to ignore files during comparison.",
    )

    args = parser.parse_args()

    old_file = args.old_file
    new_file = args.new_file
    allow_delete = args.allow_delete
    check_directory = args.check_directory
    sync_type = args.sync_type
    ignore_sheet_regex = args.ignore_sheet_regex
    ignore_file_regex = args.ignore_file_regex

    # If the paths are directories, compare all files inside them
    if check_directory:
        if not os.path.isdir(old_file) or not os.path.isdir(new_file):
            print("Both paths must be directories when using --check-directory.")
            return

        compare_directory_files(
            old_file,
            new_file,
            allow_delete,
            sync_type,
            ignore_file_regex,
            ignore_sheet_regex,
        )
    else:
        # Otherwise, compare the single files directly
        if not os.path.exists(old_file) or not os.path.exists(new_file):
            print("One or both of the provided files do not exist.")
            return

        asyncio.run(
            compare_and_sync_columns(
                old_file, new_file, allow_delete, sync_type, ignore_sheet_regex
            )
        )


if __name__ == "__main__":
    main()
