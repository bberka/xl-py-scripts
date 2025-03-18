import asyncio
import openpyxl
import aiofiles


async def read_excel(file_path):
    """
    Asynchronously reads an Excel file and returns the sheet names.
    :param file_path: str, path to the Excel file.
    :return: list of sheet names.
    """
    try:
        # Open the Excel file asynchronously
        async with aiofiles.open(file_path, mode='rb') as f:
            content = await f.read()

        # Write the content into a temporary file that openpyxl can read
        with open(file_path, 'wb') as temp_file:
            temp_file.write(content)

        # Load the Excel workbook using openpyxl
        wb = openpyxl.load_workbook(file_path)

        # Return sheet names
        return wb.sheetnames

    except Exception as e:
        print(f"Error reading the Excel file: {e}")
        return []


async def read_rows(file_path, sheet_name):
    """
    Asynchronously reads the rows of a specific sheet in the Excel file.
    :param file_path: str, path to the Excel file.
    :param sheet_name: str, name of the sheet to read.
    :return: list of rows (each row is a tuple of cell values).
    """
    try:
        # Open the Excel file asynchronously
        async with aiofiles.open(file_path, mode='rb') as f:
            content = await f.read()

        # Write the content into a temporary file that openpyxl can read
        with open(file_path, 'wb') as temp_file:
            temp_file.write(content)

        # Load the Excel workbook using openpyxl
        wb = openpyxl.load_workbook(file_path)

        # Check if the sheet exists
        if sheet_name not in wb.sheetnames:
            raise ValueError(f"Sheet '{sheet_name}' not found in the Excel file.")

        # Get the sheet and read its rows
        sheet = wb[sheet_name]
        rows = []
        for row in sheet.iter_rows(values_only=True):
            rows.append(row)

        return rows

    except Exception as e:
        print(f"Error reading the sheet '{sheet_name}': {e}")
        return []

